# 修改时间22-4-6
# 使用dp算法进行锦屏一级风光水调度
# 其中风光出力以及径流数据确定,通过调整发电水位来保证总发电量最大
# 尺度为长尺度计算,计算尺度为日尺度
# 创建人:姬新洋
import time  # 加载计时模块
import matplotlib.pyplot as plt  #加载绘图模块
import numpy as np  # 加载矩阵运算模块
import openpyxl  # 加载数据读取模块

time_start = time.time()  # 开始计时


# 根据某时段初末库用,确定该时段内平均上游水位
def h1(v1x, v2x):
    v = (v1x + v2x) / 2
    h = (-8e-6) * v ** 4 + 0.0022 * v ** 3 - 0.2104 * v ** 2 + 9.5804 * v + 1659.1
    return h


# 根据下泄流量计算下游平均水位
def h2(qx):
    h = 7e-12 * qx ** 3 - 2e-7 * qx ** 2 + 0.0034 * qx + 1633.6
    return h


# 根据上游水位计算上游库容
def v1(hx):
    v = 3.227e-6 * (hx ** 3) - 0.01535 * (hx ** 2) + (24.32 * hx) - 1.283 * (10 ** 4)
    return v


# 根据时段内入库流量,该阶段内的初末库容,计算该时段内下泄流量
def q12(v1x, v2x, qx):
    q = (qx * 24 * 3600 + (v1x - v2x) * 10 ** 8) / (24 * 3600)
    return q


# 根据上游库容计算上游水位
def h1_v(vx):
    h = (-8e-6) * vx ** 4 + 0.0022 * vx ** 3 - 0.2104 * vx ** 2 + 9.5804 * vx + 1659.1
    return h


# 根据上下游水位以及出力，反算发电流量
def q_fadian(h1x, h2x, nx):
    q = (1000 * nx) / (8.3 * (h1x - h2x))
    return q


# 根据上下游水位以及发电流量计算出力
# 同时附加流量以及出力的限制条件
def n1(qx, h1x, h2x):
    n = 0
    if qx < 0:
        n = float('-inf')  # 最小流量的限制条件
    elif qx > 2024.4:
        n = 8.3 * (h1x - h2x) * 2024.4 / 1000
        if n < 0:  # 最小出力限制条件
            n = float('-inf')
        elif n > 3600:  # 最大出力限制条件(装机容量)
            n = 3600
    else:
        n = 8.3 * (h1x - h2x) * qx / 1000
        if n < 0:  # 最小出力限制条件
            n = float('-inf')
        elif n > 3600:  # 最大出力限制条件(装机容量)
            n = 3600
    return n


# 出力计算结果单位是兆瓦


# 将水位从死水位1800m到正常蓄水位1880m离散为t份,按照库容离散
t_midu = 200
h_si = 1800
h_zhengchang = 1880
v_lisan = [0 for _ in range(t_midu)]  # 用于存储离散水位
for i in range(0, t_midu):
    v_lisan[i] = v1(h_zhengchang) - (v1(h_zhengchang) - v1(h_si)) / (t_midu - 1) * i

# 计算输入初始条件
# 需要输入的计算精度,此时为年
y_midu = 365  # 这里精度要改为365
# 需要输入起始水位
h1_start = 1800
# 需要输入结束水位
h1_finish = 1800
# 需要输入入库流量变化过程，变化过程也要替换为日变化过程
q_in = [0 for _ in range(y_midu)]
f_in = [0 for _ in range(y_midu)]
g_in = [0 for _ in range(y_midu)]
# 加载计算数据表格，并导入风光出力数据以及径流数据
wb = openpyxl.load_workbook('jisuanshuju.xlsx')
sheet1 = wb.get_sheet_by_name('入库流量（日尺度）')
sheet2 = wb.get_sheet_by_name('接入风电数据（日尺度）')
sheet3 = wb.get_sheet_by_name('接入光电数据（日尺度）')
for i in range(2, 367):
    q_in[i - 2] = sheet1.cell(row=i, column=2).value  # 导入日尺度径流数据

for i in range(2, 367):
    f_in[i - 2] = sheet2.cell(row=i, column=3).value  # 导入日尺度风出力数据

for i in range(2, 367):
    g_in[i - 2] = sheet3.cell(row=i, column=3).value  # 导入日尺度光伏出力数据

# 以下数据仅为导入数据前的测试数据，无实际用处
# q_in = [1920, 3130, 2160, 2480, 1690, 816, 506, 370, 316, 310, 384, 718]  # 较平水年入库径流变化过程
# 这里加入风光出力确定性数据
# f_in = [30, 40, 35, 40, 30, 35, 45, 30, 35, 45, 40, 35]
# g_in = [200, 180, 145, 110, 80, 90, 120, 90, 110, 160, 150, 180]

# 需要输入通道容量的最大值，这里通道容量为假设值-------------------------------------------------------------------------------
tongdao_max = 3600  #（实际的通道限制流量应该与该电站最大装机容量相同）

# 创建用于记录最优路线编号的列表
# 创建出力记录列表
na_jilu = [0 for _ in range(y_midu)]
# 创建上游水位记录列表
h1a_jilu = [0 for _ in range(y_midu)]
# 创建下泄流量记录列表
qa_jilu = [0 for _ in range(y_midu)]
# 创建下游水位记录列表
h2a_jilu = [0 for _ in range(y_midu)]
# 创建发电流量记录列表
qfadian_jilu = [0 for _ in range(y_midu)]
# 创建弃水流量记录列表
qishui_jilu = [0 for _ in range(y_midu)]
# 初始化记录第一阶段最大出力路径编号
pa_jilu = 0
# 初始化记录第三阶段最大出力路径编号
pc_jilu = 0
# 初始化记录逆推阶段的最大出力路径编号
pd_jilu = 0
# 创建记录各个阶段末水位对应的累加节点出力值记录列表
# 列表的第一位表示第几个列表,第二位表示某个列表中的第几个元素
n_max = [[0 for _ in range(200)] for _ in range(y_midu)]

# 第一阶段计算 从初水位点开始的计算
for i in range(0, t_midu):
    h1a = h1(v1(h1_start), v_lisan[i])  # 计算上游平均水位
    qa = q12(v1(h1_start), v_lisan[i], q_in[0])  # 计算下泄流量
    h2a = h2(qa)  # 下游水位计算
    na = n1(qa, h1a, h2a)  # 对应出力计算
    if abs(h1(v_lisan[i], v_lisan[i]) - h1(v1(h1_start), v1(h1_start))) > 2:  # 如果时段初末水位超过水位变幅限制，惩罚
        na = float('-inf')


    if na + f_in[0] + g_in[0] > tongdao_max:  # 进行通道约束判断
        na = tongdao_max - f_in[0] -g_in[0]  # 超出通道容量的采取限制水电出力的措施

    n_max[0][i] = na  # 将各个下游水位对应出力存入n_max中

    if na > na_jilu[0]:  # 判断是否是第一阶段的最优路径编号,如果是,就进行信息记录
        na_jilu[0] = na
        h1a_jilu[0] = h1a
        qa_jilu[0] = qa
        h2a_jilu[0] = h2a
        pa_jilu = i
        qfadian_jilu[0] = q_fadian(h1a, h2a, na)

# 第二阶段计算 动态规划遍历计算
for i in range(1, y_midu - 1):
    for j in range(0, t_midu):
        nb_max = float('-inf')  # 初始化节点出力累加值
        for k in range(0, t_midu):
            h1b = h1(v_lisan[k], v_lisan[j])  # 计算上游水位
            qb = q12(v_lisan[k], v_lisan[j], q_in[i])  # 计算下泄流量
            h2b = h2(qb)  # 计算下游水位
            nb = n1(qb, h1b, h2b)  # 出力计算
            if abs(h1(v_lisan[k], v_lisan[k]) - h1(v_lisan[j], v_lisan[j])) > 2:  # 如果时段初末水位超过水位变幅限制，惩罚
                nb = float('-inf')

            if nb + f_in[i] + g_in[i] > tongdao_max:  # 进行通道约束判断
                nb = tongdao_max - f_in[i] - g_in[i]  # 超出通道容量的采取限制水电出力的措施

            # 对应同一个阶段末节点的不同线路上的出力值,应与改阶段初位置上的节点出力值进行累加
            n_sum = n_max[i - 1][k] + nb
            if n_sum > nb_max:
                nb_max = n_sum

        n_max[i][j] = nb_max  # 将各个计算最大累加出力值存入n_max

# 第三阶段计算 到末水位点的计算
nc_max = float('-inf')  # 初始化节点出力累加值
for i in range(0, t_midu):
    h1c = h1(v_lisan[i], v1(h1_finish))  # 计算上游水位
    qc = q12(v_lisan[i], v1(h1_finish), q_in[y_midu - 1])  # 计算下泄流量
    h2c = h2(qc)  # 计算下游水位
    nc = n1(qc, h1c, h2c)  # 出力计算
    if abs(h1(v_lisan[i], v_lisan[i]) - h1(v1(h1_finish), v1(h1_finish))) > 2:  # 如果时段初末水位超过水位变幅限制，惩罚
        nc = float('-inf')

    if nc + f_in[y_midu - 1] + g_in[y_midu - 1] > tongdao_max:  # 进行通道约束判断
        nc = tongdao_max - f_in[y_midu - 1] -g_in[y_midu - 1]  # 超出通道容量的采取限制水电出力的措施

    n_max[y_midu - 1][i] = n_max[y_midu - 2][i] + nc
    if n_max[y_midu - 1][i] > nc_max:
        nc_max = n_max[y_midu - 1][i]
        pc_jilu = i

    if nc > na_jilu[y_midu - 1]:  # 记录第三阶段的最优路径,如果是,记录它
        na_jilu[y_midu - 1] = nc
        h1a_jilu[y_midu - 1] = h1c
        qa_jilu[y_midu - 1] = qc
        h2a_jilu[y_midu - 1] = h2c
        qfadian_jilu[y_midu - 1] = q_fadian(h1c, h2c, nc)

# 逆推中间阶段的最优路径,获得记录数据
pd_jilu = pc_jilu  # 将第三阶段的最优节点编号传递给逆推起点编号
for i in range(1, y_midu - 1):
    nd_max = float('-inf')  # 初始化记录值
    x_to = 0  # 初始化用于传导第二阶段的最优路径编号

    for j in range(0, t_midu):
        h1d = h1(v_lisan[j], v_lisan[pd_jilu])
        qd = q12(v_lisan[j], v_lisan[pd_jilu], q_in[y_midu - 1 - i])
        h2d = h2(qd)
        nd = n1(qd, h1d, h2d)
        # 逆推各个阶段的出力最大值,并记录最优节点编号,以便记录其他信息
        if abs(h1(v_lisan[j], v_lisan[j]) - h1(v_lisan[pd_jilu], v_lisan[pd_jilu])) > 2:  # 初末水位超过水位变幅限制，惩罚
            nd = float('-inf')

        if nd + f_in[y_midu - 1 - i] + g_in[y_midu - 1 - i] > tongdao_max:  # 进行通道约束判断
            nd = tongdao_max - f_in[y_midu - 1 - i] - g_in[y_midu - 1 - i]  # 超出通道容量的采取限制水电出力的措施

        nd_sum = n_max[y_midu - 2 - i][j] + nd
        if nd_sum > nd_max:
            nd_max = nd_sum
            x_to = j
            na_jilu[y_midu - 1 - i] = nd
            h1a_jilu[y_midu - 1 - i] = h1d
            qa_jilu[y_midu - 1 - i] = qd
            h2a_jilu[y_midu - 1 - i] = h2d
            qfadian_jilu[y_midu - 1 - i] = q_fadian(h1d, h2d, nd)  # 计算发电流量
    else:
        pd_jilu = x_to
# 结束逆推计算

time_end = time.time()  # 结束计时

time_c = time_end - time_start  # 计算计时时间
print('本次计算用时:' + str(time_c) + 's')

# 输出计算结果
n_zong = np.array(na_jilu) + np.array(f_in) + np.array(g_in)  # 风光水总出力计算
qishui_jilu = np.array(qa_jilu) - np.array(qfadian_jilu)  # 计算弃水流量过程
for i in range(y_midu):  # 由于二进制存储数据不完全等于0，故需要剔除无限接近0的数
    if qishui_jilu[i] < 0:
        qishui_jilu[i] = 0

# 打印计算过程
na_zong = list(n_zong)
print('上游水位变化过程：',h1a_jilu)
print('下游水位变化过程：',h2a_jilu)
print('下泄流量变化过程：',qa_jilu)
print('水电机组出力变化过程：',na_jilu)
print('总出力变化过程：',na_zong)
print('发电流量过程：',qfadian_jilu)
print('水库入库径流过程：',q_in)
print('弃水流量过程：',list(qishui_jilu))

# 不采用python进行结果绘制，而是将输出数据存入excel
jg_sheet = openpyxl.Workbook()
jg_sheet.create_sheet(index=0, title='上游水位变化数据')  # 创建工作表
jg_sheet.create_sheet(index=1, title='下游水位变化数据')
jg_sheet.create_sheet(index=2, title='下泄流量变化数据')
jg_sheet.create_sheet(index=3, title='出力变化数据')
jd1 = jg_sheet.get_sheet_by_name('上游水位变化数据')
jd2 = jg_sheet.get_sheet_by_name('下游水位变化数据')
jd3 = jg_sheet.get_sheet_by_name('下泄流量变化数据')
jd4 = jg_sheet.get_sheet_by_name('出力变化数据')
jd1.cell(row=1, column=1).value = '上游水位(m)'
jd2.cell(row=1, column=1).value = '下游水位(m)'
jd3.cell(row=1, column=1).value = '下泄流量(立方米每秒)'
jd3.cell(row=1, column=2).value = '机组发电流量'
jd3.cell(row=1, column=3).value = '弃水流量'
jd4.cell(row=1, column=1).value = '总出力(MW)'
jd4.cell(row=1, column=2).value = '水电机组出力(MW)'
jd4.cell(row=1, column=3).value = '风电机组出力(MW)'
jd4.cell(row=1, column=4).value = '光电机组出力(MW)'
for i in range(365):
    jd1.cell(row=i + 2, column=1).value = h1a_jilu[i]
    jd2.cell(row=i + 2, column=1).value = h2a_jilu[i]
    jd3.cell(row=i + 2, column=1).value = qa_jilu[i]
    jd3.cell(row=i + 2, column=2).value = qfadian_jilu[i]
    jd3.cell(row=i + 2, column=3).value = qishui_jilu[i]
    jd4.cell(row=i + 2, column=1).value = na_zong[i]
    jd4.cell(row=i + 2, column=2).value = na_jilu[i]
    jd4.cell(row=i + 2, column=3).value = f_in[i]
    jd4.cell(row=i + 2, column=4).value = g_in[i]

print(jg_sheet.sheetnames)
jg_sheet.save('22-2-8计算结果.xlsx')  # 保存结果

# 进行结果绘制
# 进行x轴坐标值的设置
mounth = [0 for _ in range(365)]
wb = openpyxl.load_workbook('jisuanshuju.xlsx')
sheet4 = wb.get_sheet_by_name('横坐标')
for i in range(2, 367):
    mounth[i - 2] = sheet4.cell(row=i, column=1).value

# 进行绘图中坐标点x坐标的设置
x = np.arange(0,365)+1
x[0] = 1

# 设置图形大小
plt.figure(1, figsize=(16, 8))  # 设置图形的像素大小
ax1 = plt.subplot(2, 2, 1)  # 进行子图行列划分以及位置确定
ax2 = plt.subplot(2, 2, 2)
ax3 = plt.subplot(2, 2, 3)
ax4 = plt.subplot(2, 2, 4)
plt.subplots_adjust(wspace=0.3, hspace=0.3)  # 调整子图之间的横向以及纵向间距

# 绘制上游水位变化曲线
plt.sca(ax1)  # 选定位置1处的子图
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来显示正常中文标签
plt.plot(x, h1a_jilu, label='上游水位', color='black', linewidth=2)  # 进行曲线的设置
plt.xticks(x, mounth)  # 进行x轴标记的设置

# 进行xy轴标签以及标题的设置
plt.xlabel('月份')
plt.ylabel('上游水位（m）')
plt.title('上游水位变化曲线')
plt.legend()  # 添加图例标签

# 绘制下游水位变化曲线
plt.sca(ax2)
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来显示正常中文标签
plt.plot(x, h2a_jilu, label='下游水位', color='black', linewidth=2)
plt.xticks(x, mounth)

plt.xlabel('月份')
plt.ylabel('下游水位（m）')
plt.title('下游水位变化曲线')
plt.legend()

# 绘制下泄流量变化曲线
plt.sca(ax3)
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来显示正常中文标签
plt.plot(x, qa_jilu, label='下泄流量', color='black', linewidth=2)
plt.plot(x, qfadian_jilu, label='机组发电流量', color='blue', linewidth=2)
plt.plot(x, qishui_jilu, label='弃水流量', color='red', linewidth=2)
plt.xticks(x, mounth)

plt.xlabel('月份')
plt.ylabel('下泄流量（m^3/s）')
plt.title('下泄流量变化曲线')
plt.legend()

# 绘制出力变化曲线
plt.sca(ax4)
plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来显示正常中文标签
plt.plot(x, na_jilu, label='水电机组出力', color='blue', linewidth=2)
plt.plot(x, n_zong, label='总出力', color='black', linewidth=2)
plt.plot(x, f_in, label='风电机组出力', color='yellow', linewidth=2)
plt.plot(x, g_in, label='光电机组出力', color='red', linewidth=2)
plt.xticks(x, mounth)

plt.xlabel('月份')
plt.ylabel('风光水机组出力（兆瓦）')
plt.title('出力变化曲线')
plt.legend()

plt.show()  # 显示图形
